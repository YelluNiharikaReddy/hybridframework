import openpyxl
def getRowcount(File,SheetName):
    workbook = openpyxl.load_workbook(File)
    Sheet = workbook.get_sheet_by_name(SheetName)
    return(Sheet.max_row)

def getcolumncount(File,SheetName):
    workbook = openpyxl.load_workbook(File)
    Sheet = workbook.get_sheet_by_name(SheetName)
    return(Sheet.max_column)

def readData(File,SheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(File)
    Sheet = workbook.get_sheet_by_name(SheetName)
    return Sheet.cell(row=rownum,column=columnno).value

def writeData(File,SheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(File)
    Sheet = workbook.get_sheet_by_name(SheetName)
    Sheet.cell(row=rownum , column=columnno).value = data
    workbook.save(File)